"""
extract_excel.py — Auto-generate MkDocs data-consumption Markdown pages
from MKC PowerBI Inventory (1).xlsx

Usage:
    python scripts/extract_excel.py

Outputs:
    docs/03_data-consumption/workspace-catalog.md
    docs/03_data-consumption/dataflow-inventory.md
    docs/03_data-consumption/source-inventory.md
"""

import sys
import os
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not found — run: pip install openpyxl")

SCRIPT_DIR = Path(__file__).parent
ROOT = SCRIPT_DIR.parent.parent  # mkc_assessment/
EXCEL_PATH = ROOT / "MKC PowerBI Inventory (1).xlsx"
OUT_DIR = SCRIPT_DIR.parent / "docs" / "03_data-consumption"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def load_sheet(wb, name):
    if name not in wb.sheetnames:
        sys.exit(f"Sheet '{name}' not found. Available: {wb.sheetnames}")
    ws = wb[name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            rows.append(dict(zip(headers, row)))
    return headers, rows


def safe(v):
    return str(v).strip() if v is not None else ""


# ---------------------------------------------------------------------------
# Sheet 1: DFW Lineage → workspace-catalog.md + dataflow-inventory.md
# ---------------------------------------------------------------------------

def gen_workspace_catalog(rows):
    # Build workspace → reports mapping
    ws_reports: dict[str, list[dict]] = {}
    for r in rows:
        ws = safe(r.get("workspace_name"))
        rname = safe(r.get("report_name"))
        dname = safe(r.get("dataset_name"))
        if ws and rname:
            ws_reports.setdefault(ws, [])
            if not any(x["report_name"] == rname for x in ws_reports[ws]):
                ws_reports[ws].append({"report_name": rname, "dataset_name": dname})

    total_reports = sum(len(v) for v in ws_reports.values())
    total_ws = len(ws_reports)

    lines = [
        "# Workspace Catalog",
        "",
        "Auto-generated from **DFW Lineage** sheet of `MKC PowerBI Inventory (1).xlsx`.",
        "",
        f"!!! info \"Coverage\"",
        f"    **{total_ws} workspaces** · **{total_reports} unique reports**",
        "",
    ]

    for ws_name in sorted(ws_reports.keys()):
        reports = ws_reports[ws_name]
        lines += [
            f"## {ws_name}",
            "",
            f"**{len(reports)} report{'s' if len(reports) != 1 else ''}**",
            "",
            "| # | Report Name | Semantic Model (Dataset) |",
            "|---|------------|--------------------------|",
        ]
        for i, r in enumerate(sorted(reports, key=lambda x: x["report_name"]), 1):
            lines.append(f"| {i} | {r['report_name']} | {r['dataset_name']} |")
        lines.append("")

    return "\n".join(lines)


def gen_dataflow_inventory(rows):
    # workspace → dataflows
    ws_flows: dict[str, set] = {}
    for r in rows:
        ws = safe(r.get("workspace_name"))
        df = safe(r.get("dataflow_name"))
        if ws and df:
            ws_flows.setdefault(ws, set()).add(df)

    total_flows = sum(len(v) for v in ws_flows.values())

    lines = [
        "# Dataflow Inventory",
        "",
        "Auto-generated from **DFW Lineage** sheet of `MKC PowerBI Inventory (1).xlsx`.",
        "",
        f"!!! info \"Coverage\"",
        f"    **{total_flows} unique dataflows** across **{len(ws_flows)} workspaces**",
        "",
        "## Summary by Workspace",
        "",
        "| Workspace | Dataflow Count | Dataflow Names |",
        "|-----------|---------------|----------------|",
    ]

    for ws_name in sorted(ws_flows.keys()):
        flows = sorted(ws_flows[ws_name])
        flow_list = "<br>".join(flows) if flows else "—"
        lines.append(f"| {ws_name} | {len(flows)} | {flow_list} |")

    lines += [
        "",
        "## Full Dataflow List",
        "",
        "| Workspace | Dataflow Name |",
        "|-----------|---------------|",
    ]
    for ws_name in sorted(ws_flows.keys()):
        for df in sorted(ws_flows[ws_name]):
            lines.append(f"| {ws_name} | {df} |")

    lines.append("")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Sheet 2: Source Mapping → source-inventory.md
# ---------------------------------------------------------------------------

def gen_source_inventory(rows):
    # Count by type and workspace
    type_counts: dict[str, int] = {}
    ws_sources: dict[str, list[dict]] = {}
    server_dbs: dict[str, set] = {}

    for r in rows:
        ws = safe(r.get("Workspace"))
        rtype = safe(r.get("Type"))
        server = safe(r.get("Server"))
        db = safe(r.get("DB"))

        type_counts[rtype] = type_counts.get(rtype, 0) + 1
        ws_sources.setdefault(ws, []).append(r)
        if server:
            server_dbs.setdefault(server, set()).add(db)

    total = len(rows)

    lines = [
        "# Source Inventory",
        "",
        "Auto-generated from **Source Mapping** sheet of `MKC PowerBI Inventory (1).xlsx`.",
        "",
        f"!!! info \"Coverage\"",
        f"    **{total} source connections** across **{len(ws_sources)} workspaces**",
        "",
        "## Source Type Breakdown",
        "",
        "| Source Type | Connections | % of Total |",
        "|-------------|-------------|-----------|",
    ]
    for t, cnt in sorted(type_counts.items(), key=lambda x: -x[1]):
        pct = cnt / total * 100
        lines.append(f"| {t} | {cnt} | {pct:.1f}% |")

    lines += [
        "",
        "## SQL Server & Database Inventory",
        "",
        "| Server | Databases |",
        "|--------|-----------|",
    ]
    for server in sorted(server_dbs.keys()):
        if server:
            dbs = ", ".join(sorted(d for d in server_dbs[server] if d))
            lines.append(f"| `{server}` | {dbs} |")

    lines += [
        "",
        "## Connections by Workspace",
        "",
    ]
    for ws_name in sorted(ws_sources.keys()):
        srcs = ws_sources[ws_name]
        lines += [
            f"### {ws_name}",
            "",
            f"**{len(srcs)} connection{'s' if len(srcs) != 1 else ''}**",
            "",
            "| Report | Type | Server | Database | Table |",
            "|--------|------|--------|----------|-------|",
        ]
        for s in sorted(srcs, key=lambda x: safe(x.get("reportname", ""))):
            lines.append(
                f"| {safe(s.get('reportname'))} "
                f"| {safe(s.get('Type'))} "
                f"| {safe(s.get('Server'))} "
                f"| {safe(s.get('DB'))} "
                f"| {safe(s.get('Final Table'))} |"
            )
        lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if not EXCEL_PATH.exists():
        sys.exit(f"Excel file not found: {EXCEL_PATH}")

    print(f"Loading {EXCEL_PATH.name}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    _, dfw_rows = load_sheet(wb, "DFW Lineage")
    _, src_rows = load_sheet(wb, "Source Mapping")

    out_dir = OUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    pages = {
        "workspace-catalog.md": gen_workspace_catalog(dfw_rows),
        "dataflow-inventory.md": gen_dataflow_inventory(dfw_rows),
        "source-inventory.md": gen_source_inventory(src_rows),
    }

    for fname, content in pages.items():
        path = out_dir / fname
        path.write_text(content, encoding="utf-8")
        print(f"  Written: {path.relative_to(ROOT)}")

    print("Done.")


if __name__ == "__main__":
    main()
